#!/usr/bin/env python3
"""
Gerencia o cache de regiões geográficas (bounding boxes) com rotas e comunidades associadas.

Esta classe fornece funcionalidades para armazenar, recuperar, atualizar e remover entradas de cache
relacionadas a regiões geográficas consultadas junto ao servidor OSMR. Cada entrada de cache pode conter
rotas e polilinhas de comunidades associadas, armazenadas em cache secundário específico.

Funcionalidades principais:
- Geração de chave hash para identificar regiões geográficas.
- Armazenamento persistente do cache em disco com compactação (gzip + pickle).
- Cache auxiliar para rotas (`RouteCache`) e comunidades (`PolylineCache`).
- Mecanismo debounce para gravações automáticas em segundo plano.
- Exportação do cache para planilha Excel compactada em .zip.
- Limpeza automática de entradas antigas com critérios configuráveis.

Atributos:
    cache (dict): Dicionário contendo os metadados do cache principal.
    ultimaregiao (any): Última região consultada.
    cache_file (Path): Caminho do arquivo de cache persistente no disco.
    _save_timer (threading.Timer): Temporizador de debounce para salvar o cache.
    _debounce_delay (int): Tempo em segundos antes do salvamento automático.
    _lock (threading.Lock): Lock de proteção para acessos concorrentes.

Métodos principais:
    new(regioes, diretorio): Cria nova entrada de cache.
    get_cache(regioes): Retorna o diretório do cache para a região.
    delete(regioes): Remove uma entrada de cache do disco e da memória.
    clear_cache(): Remove todas as entradas de cache.
    get_comunidades(...) / set_comunidades(...): Interface para cache de comunidades.
    exportar_cache_para_xlsx_zip(path): Exporta conteúdo do cache para planilha compactada.
    clean_old_cache_entries(meses, minimo_regioes): Remove entradas antigas com base em tempo e quantidade mínima.

Exemplos de uso:
    c = CacheBoundingBox()
    c.new({'norte': -22.1, 'sul': -22.2, 'leste': -43.1, 'oeste': -43.2}, 'regiao-abc')
    dir = c.get_cache({'norte': -22.1, 'sul': -22.2, 'leste': -43.1, 'oeste': -43.2'})
    c.exportar_cache_para_xlsx_zip(Path("/tmp/cache.zip"))
"""


import hashlib
import json
import os
import shutil
import gzip
from datetime import datetime
import threading
from pathlib import Path
import project_folders as pf
import pickle
from datetime import datetime, timedelta
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
import zipfile
import os
import atexit
import signal
import threading
import time

import regions as rg



# ---------------------------------------------------------------------------------------------------------------
class CacheBoundingBox: 
    def __init__(self):
        # Instância global do cache de rotas já pedidas ao servidor OSMR 
        
        self.cache = {}
        self.ultimaregiao = None
        self.cache_file = Path(pf.OSMR_PATH_CACHE_DATA) / "cache_boundingbox.bin.gz"
        self._save_timer = None
        self._debounce_delay = 4  # segundos
        self._lock = threading.Lock()
        self._load_from_disk_sync()
        self.gr = "-"
        self.state = "-"


    def remover_item_disco_old(self, caminho):
        if os.path.isdir(caminho):
            shutil.rmtree(caminho)
        elif os.path.isfile(caminho):
            os.remove(caminho)

    def remover_item_disco(self, caminho: Path, tentativas=80, espera=1.0):
        """
        Remove o diretório 'caminho' com tratamento de erros e reintentos.
        Ignora se o caminho não existir. Tenta novamente se estiver em uso.
        """
        if not caminho.exists():
            # wr.wLog(f"Caminho não encontrado para remoção: {caminho}", level="debug")
            return
        for tentativa in range(1, tentativas + 1):
            try:
                shutil.rmtree(caminho)
                # wr.wLog(f"Remoção bem-sucedida: {caminho}", level="debug")
                return
            except PermissionError as e:
                # wr.wLog(f"Tentativa {tentativa}: Permissão negada ao remover {caminho} - {e}", level="warning")
                time.sleep(espera)
            except Exception as e:
                #  wr.wLog(f"Tentativa {tentativa}: Erro ao remover {caminho} - {e}", level="error")
                time.sleep(espera)
        # print(f"Falha final ao remover diretório após {tentativas} tentativas: {caminho}")

    def new(self, regioes, diretorio, inforegiao="", km2_região=0):
        chave = self._hash_bbox(regioes)
        self.ultimaregiao = regioes
        now = datetime.now().strftime("%d/%m/%Y %H:%M:%S")

        # Gera uma string legível e truncada da região
        regiao_legivel = json.dumps(regioes)
        regiao_legivel = regiao_legivel.replace("\n", "").replace("\r", "")
        if len(regiao_legivel) > 2400:
            regiao_legivel = regiao_legivel[:2397] + "..."

        inforegiao = json.dumps(inforegiao).replace("\n", "").replace("\r", "")
        self.cache[chave] = {
            "regiao": regiao_legivel,
            "regiaodados": regioes,
            "gr": self.gr,
            "state": self.state,
            "diretorio": diretorio,
            "inforegiao": inforegiao,
            "km2_região": km2_região,
            "created": now,
            "lastrequest": now,
        }
        self._save_to_disk_sync()

    def chave(self, regioes):
        return self._hash_bbox(regioes)

    def _hash_bbox(self, regioes, tamanho=12):
        regioes_json = json.dumps(regioes, sort_keys=True)
        hash_obj = hashlib.sha256(regioes_json.encode("utf-8"))
        hash_completo = hash_obj.hexdigest()
        return hash_completo[:tamanho]

    def find_smallest_containing_region(self, regiao_alvo: list) -> list | None:
        """
        Finds the smallest (by area in km²) cached region that fully contains the given target region.

        Args:
            regiao_alvo (list): The target region to check containment for.

        Returns:
            list | None: The smallest region that contains the target, or None if none found.
        """
        smallest_region = None  # Initialize the variable to hold the best match (smallest containing region)
        smallest_area = float(
            "inf"
        )  # Initialize with infinity to ensure any real area is smaller

        for data in self.cache.values():  # Iterate through all cached regions
            candidate_region = data.get(
                "regiaodados"
            )  # Extract the candidate region data
            km2 = data.get(
                "km2_região", float("inf")
            )  # Get the area in km², or infinity if not available

            # Check if the target region is fully inside the candidate region
            if candidate_region and rg.is_region_inside_another(
                regiao_alvo, candidate_region
            ):
                # Update if this candidate has a smaller area than the current smallest
                if km2 < smallest_area:
                    smallest_area = km2
                    smallest_region = candidate_region

        # If the found region has the same exlusion regions return smallest_region
        # if smallest_region != None and rg.compare_regions_without_bounding_box(regiao_alvo, smallest_region):
        #     return smallest_region

        return smallest_region

    def get_cache(self, regioes):
        chave = self._hash_bbox(regioes)
        self.ultimaregiao = regioes
        entry = self.cache.get(chave)
        if entry:
            # Atualiza o timestamp de último acesso
            entry["lastrequest"] = datetime.now().strftime("%d/%m/%Y %H:%M:%S")
            self.gr = entry["gr"]
            self.state = entry["state"]
            self._schedule_save()
            return entry["regiaodados"]
        # melhorar verificando as áreas de exclusão
        regtemp = self.find_smallest_containing_region(regioes)
        if regtemp != None:
            self.ultimaregiao = regtemp
            entry = self.cache.get(self._hash_bbox(regtemp))
            self.gr = entry["gr"]
            self.state = entry["state"]
            return regtemp

        return None

    def lastrequestupdate(self, regioes):
        chave = self._hash_bbox(regioes)
        entry = self.cache.get(chave)
        if entry:
            # Atualiza o timestamp de último acesso
            entry["lastrequest"] = datetime.now().strftime("%d/%m/%Y %H:%M:%S")
            self._schedule_save()
            return entry["diretorio"]




    def delete(self, regioes):
        self.ultimaregiao = regioes
        chave = self._hash_bbox(regioes)
        dir = self.cache.get(chave, None)["diretorio"]
        if dir:
            self.remover_item_disco(Path(pf.OSMR_PATH_CACHE_DATA) / dir)
        if chave in self.cache:
            del self.cache[chave]
            self._save_to_disk_sync()
            return True
        return False

    def clear_cache(self):
        self.cache.clear()
        if self.cache_file.exists():
            self.cache_file.unlink()

    def _save_in_thread(self):
        """Salvamento real em thread separada."""
        with self._lock:
            thread = threading.Thread(target=self._save_to_disk_sync)
            thread.daemon = True
            thread.start()

    def _schedule_save(self):
        """Inicia ou reinicia o temporizador de salvamento com debounce, com proteção de lock."""
        with self._lock:
            if self._save_timer and self._save_timer.is_alive():
                self._save_timer.cancel()
            self._save_timer = threading.Timer(
                self._debounce_delay, self._save_in_thread
            )
            self._save_timer.daemon = True
            self._save_timer.start()


    def _save_to_disk_sync(self):
        data = {
            "cache": self.cache,
        }

        backup_file = self.cache_file.with_suffix(".bin.gz.bkp")
        try:
            # Renomeia o arquivo existente como backup, se existir
            if self.cache_file.exists():
                if backup_file.exists():
                    backup_file.unlink()  # Remove backup antigo, se houver
                self.cache_file.rename(backup_file)

            # Tenta salvar o novo arquivo
            with gzip.open(self.cache_file, "wb") as f:
                pickle.dump(data, f)

            # Exporta a planilha ZIP
            self.exportar_cache_para_xlsx_zip(
                Path(pf.OSMR_PATH_CACHE_DATA) / "cache_boundingbox.zip"
            )

            # Se tudo deu certo, remove o backup
            if backup_file.exists():
                backup_file.unlink()

        except Exception as e:
            # print(f"[ERRO] Falha ao salvar cache: {e}")

            # Se falhou e o backup existe, tenta restaurar
            if backup_file.exists():
                if self.cache_file.exists():
                    self.cache_file.unlink()
                backup_file.rename(self.cache_file)
                # print("[INFO] Cache restaurado a partir do backup.")

    def _load_from_disk_sync(self):
        def limpar_todos():
            self.cache = {}

        if not self.cache_file.exists():
            limpar_todos()
            return

        try:
            with gzip.open(self.cache_file, "rb") as f:
                data = pickle.load(f)
        except Exception as e:
            # print(f"[ERRO] Falha ao carregar cache principal: {e}")
            # Tenta o backup
            backup_file = self.cache_file.with_suffix(".bin.gz.bkp")
            if backup_file.exists():
                try:
                    with gzip.open(backup_file, "rb") as f:
                        data = pickle.load(f)
                    # print("[INFO] Cache restaurado a partir do backup.")
                    # Substitui o principal com o backup
                    if self.cache_file.exists():
                        self.cache_file.unlink()
                    backup_file.rename(self.cache_file)
                except Exception as e2:
                    # print(f"[ERRO] Falha ao carregar cache do backup: {e2}")
                    limpar_todos()
                    return
                finally:
                    if backup_file.exists():
                        backup_file.unlink()
            else:
                # print("[WARN] Backup de cache não encontrado.")
                limpar_todos()
                return

        self.cache = data.get("cache", {})

    def clean_old_cache_entries(self, meses=12, minimo_regioes=30):
        agora = datetime.now()
        limite = agora - timedelta(days=meses * 30)

        # Lista de chaves com lastrequest mais antigos que o limite
        chaves_para_remover = []
        for chave, valor in self.cache.items():
            try:
                lastreq = datetime.strptime(valor["lastrequest"], "%d/%m/%Y %H:%M:%S")
                if lastreq < limite:
                    chaves_para_remover.append((chave, lastreq))
            except Exception as e:
                print(f"Erro ao analisar data de {chave}: {e}")

        # Ordena por lastrequest mais antigo primeiro
        chaves_para_remover.sort(key=lambda x: x[1])

        # Só remove se ainda sobrarem pelo menos `minimo_regioes`
        total = len(self.cache)
        removiveis = max(0, total - minimo_regioes)
        chaves_para_remover = chaves_para_remover[:removiveis]

        for chave, _ in chaves_para_remover:
            diretorio = self.cache[chave]["diretorio"]
            self.remover_item_disco(Path(pf.OSMR_PATH_CACHE_DATA) / diretorio)
            del self.cache[chave]

        if chaves_para_remover:
            self._save_to_disk_sync()
            # wLog(f"{len(chaves_para_remover)} regiões antigas removidas do cache.",level="debug")
        else:
            # wLog("Nenhuma região antiga removida.",level="debug")
            pass

    def exportar_cache_para_xlsx_zip(self, caminho_zip):
        xlsx_path = caminho_zip.with_suffix(".xlsx")
        wb = Workbook()
        ws1 = wb.active
        ws1.title = "Cache Bounding Box"

        # Aba 1: Cache Bounding Box
        headers1 = [
            "Chave",
            "Regiao",
            "GR",
            "Estado",
            "Municipios Cobertos",
            "Diretório",
            "Km2 Região",
            "Criado em",
            "Último Acesso",
        ]
        ws1.append(headers1)
        col_widths1 = [len(h) for h in headers1]

        for chave, dados in self.cache.items():
            row = [
                chave,
                dados.get("regiao", ""),
                dados.get("gr", ""),
                dados.get("state", ""),
                dados.get("inforegiao", "").encode("utf-8").decode("unicode_escape"),
                dados.get("diretorio", ""),
                dados.get("km2_região", ""),
                dados.get("created", ""),
                dados.get("lastrequest", ""),
            ]
            ws1.append(row)
            for i, cell_value in enumerate(row):
                col_widths1[i] = max(col_widths1[i], len(str(cell_value)))
        for i, width in enumerate(col_widths1, 1):
            adjusted_width = min(width + 2, 50)
            ws1.column_dimensions[get_column_letter(i)].width = adjusted_width

        # Salva e compacta
        wb.save(xlsx_path)
        with zipfile.ZipFile(caminho_zip, "w", zipfile.ZIP_DEFLATED) as zipf:
            zipf.write(xlsx_path, arcname=os.path.basename(xlsx_path))
        os.remove(xlsx_path)

    # ------------------------------------------------------------------------------------------------
    def find_server_for_this_route(self, start_lat, start_lon, end_lat, end_lon):
        """
        Verifica se ambos os pontos estão dentro do bounding box da região "boundingBoxRegion"
        presente em algum item do cache.
        Retorna a primeira região correspondente (string JSON original), ou None.
        """

        # Converte entradas para float
        try:
            start_lat = float(start_lat)
            start_lon = float(start_lon)
            end_lat = float(end_lat)
            end_lon = float(end_lon)
        except ValueError:
            return None,None  # coordenadas inválidas

        for chave, dados in self.cache.items():
            chaveBuf = chave
            locregiao = dados.get("regiaodados", "")


            regioes = locregiao

            # Procura pela região "boundingBoxRegion"
            box = None
            for regiao in regioes:
                if regiao.get("name") == "boundingBoxRegion":
                    box = regiao.get("coord")
                    break

            if not box or len(box) != 4:
                continue  # bounding box ausente ou malformada

            # Extrai extremos do retângulo e converte para float
            try:
                latitudes = [float(p[0]) for p in box]
                longitudes = [float(p[1]) for p in box]
            except (ValueError, TypeError):
                continue  # ignora se houver dado inválido

            lat_min, lat_max = min(latitudes), max(latitudes)
            lon_min, lon_max = min(longitudes), max(longitudes)

            # Verifica se os dois pontos estão dentro do retângulo
            if (
                lat_min <= start_lat <= lat_max
                and lon_min <= start_lon <= lon_max
                and lat_min <= end_lat <= lat_max
                and lon_min <= end_lon <= lon_max
            ):
                return chaveBuf, locregiao

        return None,None

    # ------------------------------------------------------------------------------------------------

    def list_servers_online(self):

        return


# ---------------------------------------------------------------------------------------------------------------

# Instância global
cCacheBoundingBox = CacheBoundingBox()
cCacheBoundingBox.clean_old_cache_entries()


def shutdown():
    print("Salvando cache antes de sair...")
    cCacheBoundingBox._save_to_disk_sync()
    print("Cache salvo com sucesso.")


# Registro para finalização normal (Linux e Windows)
atexit.register(shutdown)


# Registro para interrupções (Ctrl+C, SIGTERM)
def signal_handler(sig, frame):
    shutdown()
    exit(0)


signal.signal(signal.SIGINT, signal_handler)
signal.signal(signal.SIGTERM, signal_handler)
