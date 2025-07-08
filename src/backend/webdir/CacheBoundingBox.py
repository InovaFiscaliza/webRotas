#!/usr/bin/env python3
"""
Gerencia o cache de regiões geográficas (bounding boxes) com rotas e comunidades associadas.

Esta classe fornece funcionalidades para armazenar, recuperar, atualizar e remover entradas de cache
relacionadas a regiões geográficas consultadas junto ao servidor OSMR. Cada entrada de cache pode conter
rotas e polilinhas de comunidades associadas, armazenadas em cache secundário específico.

Funcionalidades principais:
- Geração de chave hash para identificar regiões geográficas.
- Armazenamento persistente do cache em disco com compactação (gzip + pickle).
- Cache auxiliar para rotas (`RouteCache`) e comunidades (`PolylineCache`).
- Mecanismo debounce para gravações automáticas em segundo plano.
- Exportação do cache para planilha Excel compactada em .zip.
- Limpeza automática de entradas antigas com critérios configuráveis.

Atributos:
    route_cache (RouteCache): Cache de rotas por região.
    comunidades_cache (PolylineCache): Cache de polilinhas de comunidades por região.
    cache (dict): Dicionário contendo os metadados do cache principal.
    ultimaregiao (any): Última região consultada.
    cache_file (Path): Caminho do arquivo de cache persistente no disco.
    _save_timer (threading.Timer): Temporizador de debounce para salvar o cache.
    _debounce_delay (int): Tempo em segundos antes do salvamento automático.
    _lock (threading.Lock): Lock de proteção para acessos concorrentes.

Métodos principais:
    new(regioes, diretorio): Cria nova entrada de cache.
    get_cache(regioes): Retorna o diretório do cache para a região.
    delete(regioes): Remove uma entrada de cache do disco e da memória.
    clear_cache(): Remove todas as entradas de cache.
    route_cache_get(...) / set(...): Interface para cache de rotas.
    get_comunidades(...) / set_comunidades(...): Interface para cache de comunidades.
    exportar_cache_para_xlsx_zip(path): Exporta conteúdo do cache para planilha compactada.
    clean_old_cache_entries(meses, minimo_regioes): Remove entradas antigas com base em tempo e quantidade mínima.

Exemplos de uso:
    c = CacheBoundingBox()
    c.new({'norte': -22.1, 'sul': -22.2, 'leste': -43.1, 'oeste': -43.2}, 'regiao-abc')
    dir = c.get_cache({'norte': -22.1, 'sul': -22.2, 'leste': -43.1, 'oeste': -43.2'})
    c.exportar_cache_para_xlsx_zip(Path("/tmp/cache.zip"))
"""


import hashlib
import json
import os
import shutil
import gzip
from datetime import datetime
import threading
from pathlib import Path
import project_folders as pf
import pickle
from datetime import datetime, timedelta
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
import zipfile
import os
import route_cache as rc
import PolylineCache as pl
import atexit
import signal
import threading
import time 

import regions as rg


# ---------------------------------------------------------------------------------------------------------------
class CacheBoundingBox:
    def __init__(self):
        # Instância global do cache de rotas já pedidas ao servidor OSMR
        # wLog(f"Cache de regioes carregado")
        self.route_cache = rc.RouteCache()
        self.comunidades_cache = pl.PolylineCache()
        self.areas_urbanas = pl.PolylineCache()
        self.cache = {}
        self.ultimaregiao = None
        self.cache_file = Path(pf.OSMR_PATH_CACHE_DATA) / "cache_boundingbox.bin.gz"
        self._save_timer = None
        self._debounce_delay = 4  # segundos
        self._lock = threading.Lock()
        self._load_from_disk_sync()
        self.gr = "-"
        self.state = "-"

    def get_comunidades(self, regioes):
        self.lastrequestupdate(self.ultimaregiao)
        return self.comunidades_cache.get_polylines(regioes)

    def set_comunidades(self, regioes, polylinesComunidades):
        self.lastrequestupdate(self.ultimaregiao)
        self.comunidades_cache.add_polyline(regioes, polylinesComunidades)

    def remover_item_disco_old(self, caminho):
        if os.path.isdir(caminho):
            shutil.rmtree(caminho)
        elif os.path.isfile(caminho):
            os.remove(caminho)
        
    def remover_item_disco(self, caminho: Path, tentativas=80, espera=1.0):
        """
        Remove o diretório 'caminho' com tratamento de erros e reintentos.
        Ignora se o caminho não existir. Tenta novamente se estiver em uso.
        """  
        if not caminho.exists():
            # wr.wLog(f"Caminho não encontrado para remoção: {caminho}", level="debug")
            return
        for tentativa in range(1, tentativas + 1):
            try:
                shutil.rmtree(caminho)
                # wr.wLog(f"Remoção bem-sucedida: {caminho}", level="debug")
                return
            except PermissionError as e:
                # wr.wLog(f"Tentativa {tentativa}: Permissão negada ao remover {caminho} - {e}", level="warning")
                time.sleep(espera)
            except Exception as e:
                #  wr.wLog(f"Tentativa {tentativa}: Erro ao remover {caminho} - {e}", level="error")
                time.sleep(espera)
        # print(f"Falha final ao remover diretório após {tentativas} tentativas: {caminho}")            
                

    def new(self, regioes, diretorio, inforegiao="", km2_região=0):
        chave = self._hash_bbox(regioes)
        self.ultimaregiao = regioes
        now = datetime.now().strftime("%d/%m/%Y %H:%M:%S")

        # Gera uma string legível e truncada da região
        regiao_legivel = json.dumps(regioes)
        regiao_legivel = regiao_legivel.replace("\n", "").replace("\r", "")
        if len(regiao_legivel) > 2400:
            regiao_legivel = regiao_legivel[:2397] + "..."

        inforegiao = json.dumps(inforegiao).replace("\n", "").replace("\r", "")
        self.cache[chave] = {
            "regiao": regiao_legivel,
            "regiaodados": regioes,
            "gr": self.gr,
            "state": self.state,
            "diretorio": diretorio,
            "inforegiao": inforegiao,
            "km2_região": km2_região,
            "created": now,
            "lastrequest": now,
        }
        self._save_to_disk_sync()

    def chave(self, regioes):
        return self._hash_bbox(regioes)

    def _hash_bbox(self, regioes, tamanho=12):
        regioes_json = json.dumps(regioes, sort_keys=True)
        hash_obj = hashlib.sha256(regioes_json.encode("utf-8"))
        hash_completo = hash_obj.hexdigest()
        return hash_completo[:tamanho]

    def find_smallest_containing_region(self, regiao_alvo: list) -> list | None:
        """
        Finds the smallest (by area in km²) cached region that fully contains the given target region.

        Args:
            regiao_alvo (list): The target region to check containment for.

        Returns:
            list | None: The smallest region that contains the target, or None if none found.
        """
        smallest_region = None  # Initialize the variable to hold the best match (smallest containing region)
        smallest_area = float(
            "inf"
        )  # Initialize with infinity to ensure any real area is smaller

        for data in self.cache.values():  # Iterate through all cached regions
            candidate_region = data.get(
                "regiaodados"
            )  # Extract the candidate region data
            km2 = data.get(
                "km2_região", float("inf")
            )  # Get the area in km², or infinity if not available

            # Check if the target region is fully inside the candidate region
            if candidate_region and rg.is_region_inside_another(
                regiao_alvo, candidate_region
            ):
                # Update if this candidate has a smaller area than the current smallest
                if km2 < smallest_area:
                    smallest_area = km2
                    smallest_region = candidate_region

        # If the found region has the same exlusion regions return smallest_region
        # if smallest_region != None and rg.compare_regions_without_bounding_box(regiao_alvo, smallest_region):
        #     return smallest_region

        return smallest_region

    def get_cache(self, regioes):
        chave = self._hash_bbox(regioes)
        self.ultimaregiao = regioes
        entry = self.cache.get(chave)
        if entry:
            # Atualiza o timestamp de último acesso
            entry["lastrequest"] = datetime.now().strftime("%d/%m/%Y %H:%M:%S")
            self.gr = entry["gr"]
            self.state = entry["state"]
            self._schedule_save()
            return entry["regiaodados"]
        # melhorar verificando as áreas de exclusão
        regtemp = self.find_smallest_containing_region(regioes)
        if regtemp != None:
            self.ultimaregiao = regtemp
            entry = self.cache.get(self._hash_bbox(regtemp))
            self.gr = entry["gr"]
            self.state = entry["state"]            
            return regtemp

        return None

    def lastrequestupdate(self, regioes):
        chave = self._hash_bbox(regioes)
        entry = self.cache.get(chave)
        if entry:
            # Atualiza o timestamp de último acesso
            entry["lastrequest"] = datetime.now().strftime("%d/%m/%Y %H:%M:%S")
            self._schedule_save()
            return entry["diretorio"]

    def route_cache_get(self, start_lat, start_lon, end_lat, end_lon):
        self.lastrequestupdate(self.ultimaregiao)
        return self.route_cache.get(
            self.ultimaregiao, start_lat, start_lon, end_lat, end_lon
        )

    def route_cache_set(self, start_lat, start_lon, end_lat, end_lon, value):
        self.lastrequestupdate(self.ultimaregiao)
        self.route_cache.set(
            self.ultimaregiao, start_lat, start_lon, end_lat, end_lon, value
        )

    def route_cache_clear_regioes(self):
        self.route_cache.clear_regioes(self.ultimaregiao)

    def delete(self, regioes):
        self.ultimaregiao = regioes
        self.route_cache_clear_regioes()
        # self.comunidades_cache.clear_regiao(regioes)
        chave = self._hash_bbox(regioes)
        dir = self.cache.get(chave, None)["diretorio"]
        if dir:
            self.remover_item_disco(Path(pf.OSMR_PATH_CACHE_DATA) / dir)
        if chave in self.cache:
            del self.cache[chave]
            self._save_to_disk_sync()
            return True
        return False

    def clear_cache(self):
        self.cache.clear()
        if self.cache_file.exists():
            self.cache_file.unlink()

    def _save_in_thread(self):
        """Salvamento real em thread separada."""
        with self._lock:
            thread = threading.Thread(target=self._save_to_disk_sync)
            thread.daemon = True
            thread.start()

    def _schedule_save(self):
        """Inicia ou reinicia o temporizador de salvamento com debounce, com proteção de lock."""
        with self._lock:
            if self._save_timer and self._save_timer.is_alive():
                self._save_timer.cancel()
            self._save_timer = threading.Timer(
                self._debounce_delay, self._save_in_thread
            )
            self._save_timer.daemon = True
            self._save_timer.start()

    def salvar_route_cache_item(self, chave, dados):
        diretorio_rel = dados.get("diretorio")
        if not diretorio_rel:
            # print(f"[AVISO] Chave '{chave}' não possui diretório definido. Pulando...")
            return

        diretorio = Path(pf.OSMR_PATH_CACHE_DATA) / diretorio_rel
        diretorio.mkdir(parents=True, exist_ok=True)

        route_data = self.route_cache.cache.get(chave)
        if route_data is None:
            # print(f"[AVISO] Chave '{chave}' não possui dados em route_cache. Pulando...")
            return

        caminho_arquivo = diretorio / "route_cache.bin.gz"
        caminho_old = diretorio / "route_cache.bkp.gz"

        try:
            if caminho_arquivo.exists():
                caminho_old.unlink(missing_ok=True)  # Remove .old se já existir
                caminho_arquivo.rename(caminho_old)  # Renomeia para .old

            with gzip.open(caminho_arquivo, "wb") as f:
                pickle.dump(route_data, f)
            caminho_old.unlink(missing_ok=True)
            # print(f"[OK] Route cache salvo para '{chave}' em '{caminho_arquivo}'")
        except Exception as e:
            print(f"[ERRO] Falha ao salvar route_cache para '{chave}': {e}")


    def salvar_route_cache_individual(self):
        """
        Salva cada entrada do self.route_cache em um arquivo .bin.gz (Gzip + Pickle)
        dentro do diretório correspondente definido em self.cache.
        Antes de salvar, renomeia o arquivo antigo para route_cache.old.gz, se existir.
        """
        if self.ultimaregiao is not None:
            chave = self._hash_bbox(self.ultimaregiao)
            try:
                dados = self.cache[chave]
                self.salvar_route_cache_item(chave, dados)
            except KeyError:
                # wr.wLog(f"[AVISO] Chave '{chave}' não encontrada no cache. Nenhum dado salvo para esta região.", level="warning")
                pass
            except Exception as e:
                # wr.wLog(f"[ERRO] Falha ao salvar cache da chave '{chave}': {e}", level="error")
                pass
            return

        for chave, dados in self.cache.items():
            try:
                self.salvar_route_cache_item(chave, dados)
            except Exception as e:
                # wr.wLog(f"[ERRO] Falha ao salvar item do cache chave='{chave}': {e}", level="error")
                pass


    def carregar_route_cache_individual(self):
        """
        Carrega os arquivos .bin.gz (Gzip + Pickle) de cada diretório indicado em self.cache,
        restaurando os dados em self.route_cache.cache[chave].
        Em caso de erro ao carregar o arquivo principal, tenta carregar o arquivo de backup (.old.gz).
        """
        # print(f"Cache contém {len(self.cache)} entradas")
        for chave, dados in self.cache.items():
            diretorio = Path(pf.OSMR_PATH_CACHE_DATA) / dados.get("diretorio")

            if not diretorio:
                # print(f"[AVISO] Chave '{chave}' não possui diretório definido. Pulando...")
                continue

            caminho_arquivo = diretorio / "route_cache.bin.gz"
            caminho_old = diretorio / "route_cache.bkp.gz"

            if not caminho_arquivo.exists() and not caminho_old.exists():
                # print(f"[AVISO] Nenhum arquivo encontrado para chave '{chave}'")
                continue

            try:
                with gzip.open(caminho_arquivo, "rb") as f:
                    route_data = pickle.load(f)
                    self.route_cache.cache[chave] = route_data
                # print(f"[OK] Route cache carregado para '{chave}' de '{caminho_arquivo}'")
            except Exception as e:
                print(f"[ERRO] Falha ao carregar '{caminho_arquivo}': {e}")
                if caminho_old.exists():
                    try:
                        with gzip.open(caminho_old, "rb") as f:
                            route_data = pickle.load(f)
                            self.route_cache.cache[chave] = route_data
                        caminho_old.unlink(missing_ok=True)
                        print(
                            f"[RECUPERADO] Route cache carregado de backup '{caminho_old}' para '{chave}'"
                        )
                    except Exception as e2:
                        print(
                            f"[ERRO] Também falhou ao carregar backup '{caminho_old}': {e2}"
                        )
                else:
                    print(f"[AVISO] Arquivo de backup '{caminho_old}' não existe.")

    def _save_to_disk_sync(self):
        data = {
            "cache": self.cache,
            # 'route_cache': self.route_cache,
            "comunidades_cache": self.comunidades_cache,
            "areas_urbanas": self.areas_urbanas,
        }

        self.salvar_route_cache_individual()
        backup_file = self.cache_file.with_suffix(".bin.gz.bkp")
        try:
            # Renomeia o arquivo existente como backup, se existir
            if self.cache_file.exists():
                if backup_file.exists():
                    backup_file.unlink()  # Remove backup antigo, se houver
                self.cache_file.rename(backup_file)

            # Tenta salvar o novo arquivo
            with gzip.open(self.cache_file, "wb") as f:
                pickle.dump(data, f)

            # Exporta a planilha ZIP
            self.exportar_cache_para_xlsx_zip(
                Path(pf.OSMR_PATH_CACHE_DATA) / "cache_boundingbox.zip"
            )

            # Se tudo deu certo, remove o backup
            if backup_file.exists():
                backup_file.unlink()

        except Exception as e:
            # print(f"[ERRO] Falha ao salvar cache: {e}")

            # Se falhou e o backup existe, tenta restaurar
            if backup_file.exists():
                if self.cache_file.exists():
                    self.cache_file.unlink()
                backup_file.rename(self.cache_file)
                # print("[INFO] Cache restaurado a partir do backup.")

    def _load_from_disk_sync(self):
        def limpar_todos():
            self.cache = {}
            self.route_cache.cache = {}
            self.comunidades_cache.clear_all()
            self.areas_urbanas.clear_all()

        if not self.cache_file.exists():
            limpar_todos()
            return

        try:
            with gzip.open(self.cache_file, "rb") as f:
                data = pickle.load(f)
        except Exception as e:
            # print(f"[ERRO] Falha ao carregar cache principal: {e}")
            # Tenta o backup
            backup_file = self.cache_file.with_suffix(".bin.gz.bkp")
            if backup_file.exists():
                try:
                    with gzip.open(backup_file, "rb") as f:
                        data = pickle.load(f)
                    # print("[INFO] Cache restaurado a partir do backup.")
                    # Substitui o principal com o backup
                    if self.cache_file.exists():
                        self.cache_file.unlink()
                    backup_file.rename(self.cache_file)
                except Exception as e2:
                    # print(f"[ERRO] Falha ao carregar cache do backup: {e2}")
                    limpar_todos()
                    return
                finally:
                    if backup_file.exists():
                        backup_file.unlink()
            else:
                # print("[WARN] Backup de cache não encontrado.")
                limpar_todos()
                return

        self.cache = data.get("cache", {})
        # self.route_cache = data.get('route_cache', {})
        self.comunidades_cache = data.get("comunidades_cache", {})
        self.areas_urbanas = data.get("areas_urbanas", {})
        self.carregar_route_cache_individual()

    def clean_old_cache_entries(self, meses=12, minimo_regioes=30):
        agora = datetime.now()
        limite = agora - timedelta(days=meses * 30)

        # Lista de chaves com lastrequest mais antigos que o limite
        chaves_para_remover = []
        for chave, valor in self.cache.items():
            try:
                lastreq = datetime.strptime(valor["lastrequest"], "%d/%m/%Y %H:%M:%S")
                if lastreq < limite:
                    chaves_para_remover.append((chave, lastreq))
            except Exception as e:
                print(f"Erro ao analisar data de {chave}: {e}")

        # Ordena por lastrequest mais antigo primeiro
        chaves_para_remover.sort(key=lambda x: x[1])

        # Só remove se ainda sobrarem pelo menos `minimo_regioes`
        total = len(self.cache)
        removiveis = max(0, total - minimo_regioes)
        chaves_para_remover = chaves_para_remover[:removiveis]

        for chave, _ in chaves_para_remover:
            diretorio = self.cache[chave]["diretorio"]
            self.remover_item_disco(Path(pf.OSMR_PATH_CACHE_DATA) / diretorio)
            del self.cache[chave]
            self.route_cache.clear_regioes_pela_chave(chave)
            self.comunidades_cache.clear_regioes_pela_chave(chave)

        if chaves_para_remover:
            self._save_to_disk_sync()
            # wLog(f"{len(chaves_para_remover)} regiões antigas removidas do cache.",level="debug")
        else:
            # wLog("Nenhuma região antiga removida.",level="debug")
            pass

    def exportar_cache_para_xlsx_zip(self, caminho_zip):
        xlsx_path = caminho_zip.with_suffix(".xlsx")
        wb = Workbook()
        ws1 = wb.active
        ws1.title = "Cache Bounding Box"

        # Aba 1: Cache Bounding Box
        headers1 = [
            "Chave",
            "Regiao",
            "GR",
            "Estado",
            "Municipios Cobertos",
            "Diretório",
            "Km2 Região",
            "Num Rotas Cache",
            "Cache Comunidades",
            "Criado em",
            "Último Acesso",
        ]
        ws1.append(headers1)
        col_widths1 = [len(h) for h in headers1]

        for chave, dados in self.cache.items():
            numrotascached = len(self.route_cache.cache.get(chave, {}))
            numcomunidadescached = len(self.comunidades_cache.get_by_key(chave))
            row = [
                chave,
                dados.get("regiao", ""),
                dados.get("gr", ""),
                dados.get("state", ""),
                dados.get("inforegiao", "").encode("utf-8").decode("unicode_escape"),
                dados.get("diretorio", ""),
                dados.get("km2_região", ""),
                str(numrotascached),
                str(numcomunidadescached),
                dados.get("created", ""),
                dados.get("lastrequest", ""),
            ]
            ws1.append(row)
            for i, cell_value in enumerate(row):
                col_widths1[i] = max(col_widths1[i], len(str(cell_value)))
        for i, width in enumerate(col_widths1, 1):
            adjusted_width = min(width + 2, 50)
            ws1.column_dimensions[get_column_letter(i)].width = adjusted_width

        # Aba 2: Cache Municipios e Áreas Urbanizadas
        ws2 = wb.create_sheet(title="Cache Municipios")
        headers2 = [
            "Cidade-UF",
            "Qtde Polígonos Município",
            "Qtde Polígonos Áreas Urbanizadas",
        ]
        ws2.append(headers2)
        col_widths2 = [len(h) for h in headers2]
        for chave_regiao, valor in self.areas_urbanas.cache.items():
            try:
                cache_polylines = valor.get("polyline", [])
                cidade_uf = cache_polylines[0]
                pol_municipio = cache_polylines[1]
                pol_urbanizadas = cache_polylines[2]
                q_pol_municipio = (
                    len(pol_municipio) if hasattr(pol_municipio, "__len__") else 1
                )
                q_pol_urbanizadas = (
                    len(pol_urbanizadas) if hasattr(pol_urbanizadas, "__len__") else 1
                )
                row = [cidade_uf, str(q_pol_municipio), str(q_pol_urbanizadas)]
                ws2.append(row)
                for i, cell_value in enumerate(row):
                    col_widths2[i] = max(col_widths2[i], len(str(cell_value)))
            except Exception as e:
                ws2.append([f"Erro em {chave_regiao}", str(e), ""])

        for i, width in enumerate(col_widths2, 1):
            adjusted_width = min(width + 2, 50)
            ws2.column_dimensions[get_column_letter(i)].width = adjusted_width

        # Salva e compacta
        wb.save(xlsx_path)
        with zipfile.ZipFile(caminho_zip, "w", zipfile.ZIP_DEFLATED) as zipf:
            zipf.write(xlsx_path, arcname=os.path.basename(xlsx_path))
        os.remove(xlsx_path)

    def find_server_for_this_route(self, start_lat, start_lon, end_lat, end_lon):
        """
        Verifica se ambos os pontos estão dentro do bounding box da região "boundingBoxRegion"
        presente em algum item do cache.
        Retorna a primeira região correspondente (string JSON original), ou None.
        """

        for dados in self.cache.values():
            locregiao = dados.get("regiao", "")

            try:
                regioes = json.loads(locregiao)
            except (json.JSONDecodeError, TypeError):
                continue  # pula se a string for inválida

            # Procura pela região "boundingBoxRegion"
            box = None
            for regiao in regioes:
                if regiao.get("name") == "boundingBoxRegion":
                    box = regiao.get("coord")
                    break

            if not box or len(box) != 4:
                continue  # bounding box ausente ou malformada

            # Extrai extremos do retângulo
            latitudes = [p[0] for p in box]
            longitudes = [p[1] for p in box]
            lat_min, lat_max = min(latitudes), max(latitudes)
            lon_min, lon_max = min(longitudes), max(longitudes)

            # Verifica se os dois pontos estão dentro do retângulo
            if (lat_min <= start_lat <= lat_max and lon_min <= start_lon <= lon_max and
                lat_min <= end_lat <= lat_max and lon_min <= end_lon <= lon_max):
                return locregiao  # ou return dados, se quiser mais contexto

        return None

    
    def list_servers_online(self):
        
        return
# ---------------------------------------------------------------------------------------------------------------

# Instância global
cCacheBoundingBox = CacheBoundingBox()
cCacheBoundingBox.clean_old_cache_entries()


def shutdown():
    print("Salvando cache antes de sair...")
    cCacheBoundingBox._save_to_disk_sync()
    print("Cache salvo com sucesso.")


# Registro para finalização normal (Linux e Windows)
atexit.register(shutdown)


# Registro para interrupções (Ctrl+C, SIGTERM)
def signal_handler(sig, frame):
    shutdown()
    exit(0)


signal.signal(signal.SIGINT, signal_handler)
signal.signal(signal.SIGTERM, signal_handler)
